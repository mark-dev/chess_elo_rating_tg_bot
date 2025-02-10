from typing import List
from datetime import datetime

from tempfile import NamedTemporaryFile

from player_info import PlayerInfo
from game_record import GameRecord
from openpyxl import Workbook, load_workbook


def players_to_excel(
        players: List[PlayerInfo]
):
    def create_service_sheet(wb: Workbook):
        ws = wb.create_sheet("Info")
        ws.title = 'Info'

        now = datetime.now()
        timestamp_str = now.strftime("%Y-%m-%d %H:%M")
        ws.cell(row=1, column=1, value="Дата создания")
        ws.cell(row=1, column=2, value=timestamp_str)

        return ws, timestamp_str

    players.sort(key=lambda x: -x.rating)

    wb = Workbook()

    ws = wb.active
    ws.title = "Rating"

    _service_sheet, timestamp_str = create_service_sheet(wb)

    header_row_index = 1

    headers = PlayerInfo.get_row_data_title()
    for header_index, h in enumerate(headers):
        cell = ws.cell(row=header_row_index, column=header_index + 1)
        cell.value = h

    for player_index, p in enumerate(players):
        row_index = header_row_index + 1 + player_index
        cell_values = p.to_excel_row_data()
        for cell_index, cell_value in enumerate(cell_values):
            cell = ws.cell(row=row_index, column=1 + cell_index)
            cell.value = cell_value
    datetime.now()
    tmp = NamedTemporaryFile(suffix=f'-player-rating-{timestamp_str}.xlsx')
    wb.save(tmp.name)
    tmp.seek(0)

    return tmp


def read_players_from_excel(file_or_stream) -> List[PlayerInfo]:
    wb = load_workbook(file_or_stream)
    ws = wb.active

    parsed_players = []

    for row in ws.iter_rows(
            min_row=2,
            max_row=10000,
            values_only=True,
            max_col=len(PlayerInfo.get_row_data_title()),
    ):
        if not any(row):  # EOF
            break

        p = PlayerInfo.from_excel_row_data(row)
        parsed_players.append(p)

    return parsed_players


def read_game_records_from_excel(file_or_stream) -> List[GameRecord]:
    wb = load_workbook(file_or_stream)
    ws = wb.active

    games = []

    for row in ws.iter_rows(
            min_row=2,
            max_row=10000,
            values_only=True,
            max_col=GameRecord.ROW_DATA_EXPECTED_LEN,
    ):
        if not any(row):  # EOF
            break
        games.append(GameRecord.from_row(row))

    return games
